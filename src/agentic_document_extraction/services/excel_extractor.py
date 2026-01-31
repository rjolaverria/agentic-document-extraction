"""Native Excel file parser for structured spreadsheet extraction."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from agentic_document_extraction.excel_document import (
    ExcelCell,
    ExcelDocument,
    ExcelSheet,
)


@dataclass
class ExcelExtractionOptions:
    """Options controlling Excel extraction behaviour."""

    sheet_name: str | None = None
    include_formulas: bool = True
    max_rows: int | None = None
    max_columns: int | None = None


class ExcelExtractor:
    """Extract structured data from Excel workbooks using openpyxl."""

    def extract_from_path(
        self, file_path: Path, options: ExcelExtractionOptions | None = None
    ) -> ExcelDocument:
        """Extract structured data from an Excel file."""
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        opts = options or ExcelExtractionOptions()
        # Load twice: once to capture formulas, once for computed values
        workbook = load_workbook(filename=file_path, data_only=False, read_only=False)
        computed_wb = load_workbook(filename=file_path, data_only=True, read_only=False)

        sheet_names = workbook.sheetnames
        active_sheet_name = opts.sheet_name or workbook.active.title
        if opts.sheet_name and opts.sheet_name not in sheet_names:
            raise ValueError(f"Sheet '{opts.sheet_name}' not found in workbook")

        target_sheet_names = [opts.sheet_name] if opts.sheet_name else sheet_names
        sheets: list[ExcelSheet] = []

        for name in target_sheet_names:
            sheet = workbook[name]
            computed_sheet = computed_wb[name]
            sheets.append(
                self._extract_sheet(
                    sheet,
                    computed_sheet,
                    include_formulas=opts.include_formulas,
                    max_rows=opts.max_rows,
                    max_columns=opts.max_columns,
                )
            )

        metadata = {
            "sheet_names": sheet_names,
            "has_formulas": any(
                self._sheet_has_formulas(workbook[name]) for name in sheet_names
            ),
        }

        return ExcelDocument(
            sheets=sheets,
            active_sheet=active_sheet_name,
            metadata=metadata,
        )

    def get_sheet_names(self, file_path: Path) -> list[str]:
        """List all sheet names in a workbook."""
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        wb = load_workbook(filename=file_path, read_only=True)
        return list(wb.sheetnames)

    def extract_as_dataframe(
        self, file_path: Path, sheet_name: str | None = None
    ) -> pd.DataFrame:
        """Extract a worksheet as a pandas DataFrame."""
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        target_sheet = sheet_name or wb.active.title
        if target_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{target_sheet}' not found in workbook")
        ws = wb[target_sheet]
        data = list(ws.values)
        # If first row is header use it, else create generic
        headers = data[0] if data else []
        rows = data[1:] if len(data) > 1 else []
        return pd.DataFrame(rows, columns=headers if headers else None)

    # ------------------------------------------------------------------ #
    # Internal helpers
    # ------------------------------------------------------------------ #

    def _extract_sheet(
        self,
        sheet: Worksheet,
        computed_sheet: Worksheet,
        *,
        include_formulas: bool,
        max_rows: int | None,
        max_columns: int | None,
    ) -> ExcelSheet:
        """Extract a single worksheet."""
        rows: list[list[ExcelCell]] = []
        row_iter: Iterable[tuple[Cell, ...]] = sheet.iter_rows(
            max_row=max_rows, max_col=max_columns
        )
        computed_iter = computed_sheet.iter_rows(
            max_row=max_rows, max_col=max_columns, values_only=True
        )

        for row_cells, computed_values in zip(row_iter, computed_iter, strict=True):
            excel_row: list[ExcelCell] = []
            for cell, computed_value in zip(row_cells, computed_values, strict=True):
                excel_row.append(
                    self._build_cell(
                        cell,
                        computed_value=computed_value,
                        include_formulas=include_formulas,
                        merged_ranges=sheet.merged_cells.ranges,
                    )
                )
            rows.append(excel_row)

        headers = self._derive_headers(rows)
        return ExcelSheet(
            name=sheet.title,
            rows=rows,
            headers=headers,
            row_count=sheet.max_row
            if max_rows is None
            else min(sheet.max_row, max_rows),
            column_count=sheet.max_column
            if max_columns is None
            else min(sheet.max_column, max_columns),
        )

    def _build_cell(
        self,
        cell: Cell,
        *,
        computed_value: Any,
        include_formulas: bool,
        merged_ranges: Any,
    ) -> ExcelCell:
        """Create ExcelCell with preserved types and formula metadata."""
        is_merged = any(cell.coordinate in rng for rng in merged_ranges)
        formula = None
        value = computed_value

        if cell.data_type == "f":
            formula = str(cell.value) if cell.value is not None else None
            if include_formulas and value is None:
                value = formula
        elif value is None:
            value = cell.value

        data_type = self._map_data_type(cell, value)

        return ExcelCell(
            value=value,
            data_type=data_type,
            formula=formula if include_formulas else None,
            is_merged=is_merged,
        )

    @staticmethod
    def _map_data_type(cell: Cell, value: Any) -> str:
        """Map openpyxl data types to human-readable types."""
        if cell.data_type == "f":
            return "formula"
        if getattr(cell, "is_date", False):
            return "date"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        if value is None:
            return "null"
        return "string"

    @staticmethod
    def _derive_headers(rows: list[list[ExcelCell]]) -> list[str] | None:
        """Derive headers from the first non-empty row if possible."""
        if not rows:
            return None
        first_row = rows[0]
        if all(cell.value is None for cell in first_row):
            return None
        return [f"col_{idx + 1}" for idx, _ in enumerate(first_row)]

    @staticmethod
    def _sheet_has_formulas(sheet: Worksheet) -> bool:
        return any(cell.data_type == "f" for row in sheet.iter_rows() for cell in row)
