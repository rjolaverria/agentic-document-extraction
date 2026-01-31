"""Tests for the native ExcelExtractor."""

from __future__ import annotations

import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from agentic_document_extraction.services.excel_extractor import (
    ExcelExtractionOptions,
    ExcelExtractor,
)


def _make_workbook() -> tuple[Workbook, Path]:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Name"
    ws1["B1"] = "Amount"
    ws1["A2"] = "Alice"
    ws1["B2"] = 123.45
    ws1["A3"] = "Bob"
    ws1["B3"] = 10
    ws1["C2"] = True
    ws1["D2"] = datetime(2024, 1, 15)
    ws1["E2"] = "=SUM(B2,B3)"
    ws1.merge_cells("A4:B4")
    ws1["A4"] = "Merged"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Secondary"
    fd = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    wb.save(fd)
    return wb, fd


def test_extract_preserves_types_and_formulas() -> None:
    _, file_path = _make_workbook()
    extractor = ExcelExtractor()

    doc = extractor.extract_from_path(file_path)
    sheet = doc.sheets[0]
    cells = sheet.rows

    # Numbers and strings
    assert cells[1][1].data_type == "number"
    assert cells[1][0].data_type == "string"

    # Boolean and date
    assert cells[1][2].data_type == "boolean"
    assert cells[1][3].data_type == "date"

    # Formula captured
    assert cells[1][4].formula is not None
    assert cells[1][4].data_type == "formula"

    # Merged cells flagged
    merged_row = cells[3]
    assert merged_row[0].is_merged is True
    assert merged_row[1].is_merged is True

    file_path.unlink()


def test_sheet_selection_and_metadata() -> None:
    _, file_path = _make_workbook()
    extractor = ExcelExtractor()

    doc = extractor.extract_from_path(
        file_path, ExcelExtractionOptions(sheet_name="Sheet2")
    )

    assert doc.active_sheet == "Sheet2"
    assert len(doc.sheets) == 1
    assert doc.sheets[0].name == "Sheet2"
    assert "Sheet1" in doc.metadata["sheet_names"]

    file_path.unlink()


def test_extract_as_dataframe() -> None:
    _, file_path = _make_workbook()
    extractor = ExcelExtractor()

    df = extractor.extract_as_dataframe(file_path, sheet_name="Sheet1")
    assert isinstance(df, pd.DataFrame)
    assert list(df.columns[:2]) == ["Name", "Amount"]
    assert df.iloc[0]["Name"] == "Alice"

    file_path.unlink()
